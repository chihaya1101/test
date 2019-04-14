#----------------------------------------------------------------------

##製造番号,検査日,確認日の一覧をfilelist.csvに作成
import os, csv, time

list_file = 'filelist.csv'
date_format = '%Y/%m/%d %H:%M:%S'

csv_file = open(list_file, 'w', newline='')
csv_writer = csv.writer(csv_file)

for filename in os.listdir():
    if os.path.isfile(filename) \
            and os.path.basename(__file__) != filename \
            and list_file != filename:

        row = []
        # ファイル名 -> 製造番号
        row.append(filename)
        # ファイル更新日時 -> 検査日
        row.append(time.strftime(date_format, \
                                 time.localtime(os.path.getmtime(filename))))
        # ファイル作成日時 -> 確認日
        row.append(time.strftime(date_format, \
                                 time.localtime(os.path.getctime(filename))))
        csv_writer.writerow(row)
csv_file.close()

#----------------------------------------------------------------------

##filelist.csvをfilelist.xlsxに変換
import pandas as pd

# CSVファイルの読み込み
data = pd.read_csv('filelist.csv')

# Excel形式で出力
data.to_excel('filelist.xlsx', encoding='utf-8', index=False, header=False)

#----------------------------------------------------------------------

##管理表.xlsxを作成する
import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.colors  import Color, YELLOW
from openpyxl.styles.fills   import PatternFill, FILL_SOLID
from openpyxl.styles         import Alignment

book = openpyxl.Workbook()

active_sheet = book.active

#カラムの作成
active_sheet['A1'] = '製造番号'
active_sheet.column_dimensions['A'].width = 12
active_sheet['B1'] = '検査日'
active_sheet.column_dimensions['B'].width = 15
active_sheet['C1'] = '確認日'
active_sheet.column_dimensions['C'].width = 15
active_sheet['D1'] = '備考'
active_sheet.column_dimensions['D'].width = 30

# 罫線(外枠)を設定
border = Border(top   =Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000'),
                left  =Side(style='thin', color='000000'),
                right =Side(style='thin', color='000000'))

# セルに罫線を設定
for row_num in range(1,2):
    for col_num in range(1,5):
        active_sheet.cell(row=row_num ,column=col_num).border = border

#セルの塗りつぶし
active_sheet['A1'].fill = PatternFill(
    patternType=FILL_SOLID,
    fgColor=openpyxl.styles.colors.Color(rgb=YELLOW))
active_sheet['B1'].fill = PatternFill(
    patternType=FILL_SOLID,
    fgColor=openpyxl.styles.colors.Color(rgb=YELLOW))
active_sheet['C1'].fill = PatternFill(
    patternType=FILL_SOLID,
    fgColor=openpyxl.styles.colors.Color(rgb=YELLOW))
active_sheet['D1'].fill = PatternFill(
    patternType=FILL_SOLID,
    fgColor=openpyxl.styles.colors.Color(rgb=YELLOW))

book.save('管理表.xlsx')

#----------------------------------------------------------------------

##filelist.xlsxと管理表.xlsxを統合する

